from flask import Flask, render_template, request, redirect, flash
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

app = Flask(__name__)
app.secret_key = 'secret123'

EXCEL_FILE = 'sales.xlsx'

if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(["التاريخ", "اسم المنتج", "الكمية", "السعر", "المجموع", "اسم العميل", "اسم الكاشير"])
    wb.save(EXCEL_FILE)

@app.route('/', methods=['GET', 'POST'])
def cashier():
    if request.method == 'POST':
        product = request.form['product']
        quantity = int(request.form['quantity'])
        price = float(request.form['price'])
        total = quantity * price
        customer = request.form.get('customer', '')
        cashier = request.form['cashier']
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append([now, product, quantity, price, total, customer, cashier])
        wb.save(EXCEL_FILE)

        flash('تم تسجيل الشراء بنجاح!')
        return redirect('/')

    return render_template("index.html")

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0')
